"""記帳 Excel 匯入/匯出。

格式：每月一個工作表，多個帳戶區塊橫向並排，每區塊 4 欄（日期、金額、分類、備註），
第一列為帳戶名稱與小計公式，另有「選項」工作表存分類清單。
"""
import io
from collections import Counter
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from .models import ExpenseAccount, ExpenseCategory, ExpenseRecord

GROUP_WIDTH = 4  # 日期、金額、分類、備註
OPTIONS_SHEET = "選項"


def _cell_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        try:
            return datetime.strptime(v.strip()[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def _text(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _parse_workbook(content: bytes):
    """解析 xlsx，回傳 (帳戶名稱列表, 記錄列表, 分類列表)。"""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    account_names: list[str] = []
    rows: list[tuple[str, date, int, str | None, str | None]] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == OPTIONS_SHEET:
            continue
        data = list(wb[sheet_name].iter_rows(values_only=True))
        if not data:
            continue
        header = data[0]
        # 偵測帳戶區塊：第一列該欄是字串、且往右第 2 欄是「分類」
        groups: list[tuple[int, str]] = []
        for c in range(0, len(header) - 2, GROUP_WIDTH):
            name = header[c]
            if isinstance(name, str) and name.strip() and header[c + 2] == "分類":
                groups.append((c, name.strip()))
                if name.strip() not in account_names:
                    account_names.append(name.strip())
        # 逐區塊處理，沿用「上一筆有效日期」補上沒填日期的沖銷/調整列，
        # 讓匯入總額與 Excel 的 SUM 一致
        for c, name in groups:
            last_date: date | None = None
            for row in data[1:]:
                if c + 1 >= len(row):
                    continue
                amount = row[c + 1]
                if not isinstance(amount, (int, float)):
                    continue
                d = _cell_date(row[c]) or last_date
                if d is None:
                    continue
                last_date = d
                category = _text(row[c + 2]) if c + 2 < len(row) else None
                note = _text(row[c + 3]) if c + 3 < len(row) else None
                rows.append((name, d, round(amount), category, note))

    option_names: list[str] = []
    if OPTIONS_SHEET in wb.sheetnames:
        for row in wb[OPTIONS_SHEET].iter_rows(values_only=True):
            t = _text(row[0]) if row else None
            if t and t not in option_names:
                option_names.append(t)

    wb.close()
    return account_names, rows, option_names


async def import_workbook(db: AsyncSession, content: bytes) -> dict:
    """可重複執行的匯入：已存在的相同記錄會略過，只補上新的。"""
    account_names, rows, option_names = _parse_workbook(content)

    result = await db.execute(select(ExpenseAccount))
    accounts = {a.name: a for a in result.scalars().all()}
    has_default = any(a.is_default for a in accounts.values())
    accounts_created = 0
    for name in account_names:
        if name not in accounts:
            acc = ExpenseAccount(name=name, is_default=not has_default)
            has_default = True
            db.add(acc)
            accounts[name] = acc
            accounts_created += 1
    await db.flush()

    result = await db.execute(select(ExpenseCategory.name))
    existing_cats = set(result.scalars().all())
    categories_created = 0
    for i, name in enumerate(option_names):
        if name not in existing_cats:
            db.add(ExpenseCategory(name=name, sort_order=i))
            categories_created += 1

    # 去重：同一組（帳戶、日期、金額、分類、備註）DB 已有幾筆就略過幾筆，
    # 允許真實的重複消費（同天同店同價）也能正確匯入
    result = await db.execute(select(ExpenseRecord))
    existing = Counter(
        (r.account_id, r.date, r.amount, r.category or "", r.note or "")
        for r in result.scalars().all()
    )
    imported = 0
    skipped = 0
    for name, d, amount, category, note in rows:
        key = (accounts[name].id, d, amount, category or "", note or "")
        if existing[key] > 0:
            existing[key] -= 1
            skipped += 1
            continue
        db.add(ExpenseRecord(
            date=d, amount=amount, category=category, note=note,
            account_id=accounts[name].id,
        ))
        imported += 1

    await db.commit()
    return {
        "accounts_created": accounts_created,
        "categories_created": categories_created,
        "imported": imported,
        "skipped": skipped,
    }


async def export_workbook(db: AsyncSession) -> bytes:
    """匯出成原 Excel 格式：每月一表、帳戶並排、選項表。"""
    result = await db.execute(select(ExpenseAccount).order_by(ExpenseAccount.id))
    accounts = list(result.scalars().all())
    result = await db.execute(
        select(ExpenseRecord).order_by(ExpenseRecord.date, ExpenseRecord.id)
    )
    records = result.scalars().all()
    result = await db.execute(
        select(ExpenseCategory).order_by(ExpenseCategory.sort_order, ExpenseCategory.id)
    )
    categories = result.scalars().all()

    # (帳戶id → 欄位群組順序)；無帳戶的記錄歸到「未指定」群組
    group_ids: list[int | None] = [a.id for a in accounts]
    group_names: list[str] = [a.name for a in accounts]
    if any(r.account_id is None for r in records):
        group_ids.append(None)
        group_names.append("未指定")

    months: dict[str, dict[int | None, list[ExpenseRecord]]] = {}
    for r in records:
        m = r.date.strftime("%Y-%m")
        months.setdefault(m, {}).setdefault(r.account_id, []).append(r)

    wb = Workbook()
    wb.remove(wb.active)
    for m in sorted(months):
        ws = wb.create_sheet(m)
        per_account = months[m]
        for gi, (gid, gname) in enumerate(zip(group_ids, group_names)):
            c0 = gi * GROUP_WIDTH + 1
            recs = per_account.get(gid, [])
            amount_col = get_column_letter(c0 + 1)
            ws.cell(row=1, column=c0, value=gname)
            ws.cell(row=1, column=c0 + 1,
                    value=f"=SUM({amount_col}2:{amount_col}{len(recs) + 1 if recs else 2})")
            ws.cell(row=1, column=c0 + 2, value="分類")
            ws.cell(row=1, column=c0 + 3, value="備註")
            for i, r in enumerate(recs):
                cell = ws.cell(row=i + 2, column=c0, value=r.date)
                cell.number_format = "yyyy-mm-dd"
                ws.cell(row=i + 2, column=c0 + 1, value=r.amount)
                ws.cell(row=i + 2, column=c0 + 2, value=r.category)
                ws.cell(row=i + 2, column=c0 + 3, value=r.note)
        sum_refs = ",".join(
            f"${get_column_letter(gi * GROUP_WIDTH + 2)}$1" for gi in range(len(group_ids))
        )
        ws.cell(row=1, column=len(group_ids) * GROUP_WIDTH + 1, value=f"=SUM({sum_refs})")

    ws_opt = wb.create_sheet(OPTIONS_SHEET)
    for i, c in enumerate(categories):
        ws_opt.cell(row=i + 1, column=1, value=c.name)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
